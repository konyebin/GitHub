#!/usr/bin/env python3
"""Build committed sample xlsx files (no symlinks to Downloads)."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.workbook.check_back_template import CHECK_BACK_HEADERS  # noqa: E402

SECTION = "Business Insight"
FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HDR_FILL = PatternFill(start_color="0D1526", end_color="0D1526", fill_type="solid")
HDR_FONT = Font(bold=True, color="94A3B8", size=10)

INSTALL_BASE_HEADERS = [
    "Account Name",
    "Collab AOV",
    "Partner Name",
    "Subscription Ref Id",
    "CSM_NAME",
    "SFDC_OPPORTUNITY_NAME",
    "Lifecycle Stage_Current",
    "Risk2_0_current",
    "Webex Calling MT Provisioned Seats",
    "Webex Calling DI Provisioned Seats",
    "Cloud Calling Billed Seats",
    "Total Billed Seats",
]

INSTALL_BASE_SAMPLE_ROW = {
    "Account Name": "Sample Account (demo)",
    "Collab AOV": 250000,
    "Partner Name": "Sample Partner",
    "Subscription Ref Id": "Sub900001",
    "CSM_NAME": "Sample CSM",
    "SFDC_OPPORTUNITY_NAME": "https://ciscocollaboration.lightning.force.com/lightning/r/Opportunity/006000000000000/view",
    "Lifecycle Stage_Current": "Adopt",
    "Risk2_0_current": "Low",
    "Webex Calling MT Provisioned Seats": 120,
    "Webex Calling DI Provisioned Seats": 30,
    "Cloud Calling Billed Seats": 500,
    "Total Billed Seats": 500,
}


def build_check_back_template(dest: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col, name in enumerate(CHECK_BACK_HEADERS, start=1):
        ws.cell(1, col, value=SECTION if col == 1 else "")
        ws.cell(1, col).fill = FILL
        ws.cell(2, col, value=name)
        ws.cell(2, col).fill = HDR_FILL
        ws.cell(2, col).font = HDR_FONT
        ws.cell(2, col).alignment = Alignment(wrap_text=True, vertical="top")
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def build_install_base_sample(dest: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Install Base"
    for col, name in enumerate(INSTALL_BASE_HEADERS, start=1):
        ws.cell(1, col, value=name)
    for col, name in enumerate(INSTALL_BASE_HEADERS, start=1):
        ws.cell(2, col, value=INSTALL_BASE_SAMPLE_ROW.get(name, ""))
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def main() -> None:
    samples = ROOT / "samples"
    build_check_back_template(samples / "check_back_template.xlsx")
    build_install_base_sample(samples / "install_base_sample.xlsx")
    print(f"Wrote {samples / 'check_back_template.xlsx'}")
    print(f"Wrote {samples / 'install_base_sample.xlsx'}")


if __name__ == "__main__":
    main()

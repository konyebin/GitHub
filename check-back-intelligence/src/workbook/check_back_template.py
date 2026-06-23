"""Load and write Check Back workbook preserving two-row header structure."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

CHECK_BACK_HEADERS = [
    "Opportunity Name",
    "Opportunity (linked)",
    "SL2",
    "TCV $",
    "Closed on (MM/YY)",
    "Competitor",
    "Migrating from",
    "Migrating to",
    "Partner",
    "CSM Engagement Model (linked)",
    "CSM name",
    "Sub #",
    "Sub Term",
    "Sub start date (MM/DD/YYYY)",
    "Add-ons included or not",
    "Calling Setup Assist included (Y/N)",
    "Provisioned/Entitled Lic Calling",
    "Active Lic Calling",
    "Customer org id",
    "Notes from Calling Analytics",
    "Notes from provisioned features",
    "CCEP trial (Y/N)",
    " (G/Y/R)",
    "TAC/BEMS",
    # PowerPoint Business Insight columns (Subscription / Provisioning / Features)
    "AAR $",
    "Collab AE/SE",
    "Service lines",
    "Sub term (months)",
    "Subscription dates",
    "Platforms",
    "Trend active users 90d",
    "Trend call volume 90d",
    "Lic Professional (used/entitled)",
    "Lic Standard (used/entitled)",
    "Lic Workspace (used/entitled)",
    "Active % of provisioned",
    "External calls",
    "Meetings usage",
    "Messaging usage",
    "Auto Attendant count",
    "Hunt Groups count",
    "Call Queues count",
    "Connected-UC (Y/N)",
    "Virtual Lines count",
    "Data gathered by",
    "Data gathered date",
    "Salesforce URL",
]


def load_template(path: str | Path) -> tuple[openpyxl.Workbook, dict[str, int]]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header_row = 2
    col_index: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value:
            col_index[str(cell.value).strip()] = col_idx
    return wb, col_index


def append_row(
    ws,
    col_index: dict[str, int],
    row_data: dict[str, Any],
    low_confidence: set[str] | None = None,
) -> int:
    """Append data row; highlight low-confidence cells."""
    low_confidence = low_confidence or set()
    row_num = ws.max_row + 1
    for field, col_idx in col_index.items():
        val = row_data.get(field, "")
        if isinstance(val, dict):
            val = val.get("value", "")
        if val in (None, ""):
            continue
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        if field in low_confidence:
            cell.fill = YELLOW
    return row_num


def flatten_merged_fields(merged: dict[str, Any]) -> tuple[dict[str, Any], set[str]]:
    """Convert {field: {value, confidence}} to flat row + low-confidence set."""
    row: dict[str, Any] = {}
    low: set[str] = set()
    for field, meta in merged.items():
        if isinstance(meta, dict):
            row[field] = meta.get("value", "")
            if meta.get("confidence") == "low":
                low.add(field)
        else:
            row[field] = meta
    return row, low

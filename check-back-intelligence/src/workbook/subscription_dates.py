"""Add and fill Subscription dates from Sub start + term or existing range text."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .powerpoint_backfill import HEADER_FONT, HEADER_FILL, SECTION_FILL, _col_for_field, _header_map
from .slide_verified import VERIFIED_SLIDE_BY_ORG

SUBSCRIPTION_DATES_COL = "Subscription dates"

# Explicit slide text where start/term columns are not enough
VERIFIED_SUBSCRIPTION_DATES: dict[str, str] = {
    "fb078f9f-4332-4ace-8b1f-6a46313aa4ea": "22-May-2025 to 21-May-2030",
    "40357664-235c-44a7-9127-3b4518eeb757": (
        "28-Apr-2025 to 27-Apr-2030 (Webex); 01-Jul-2024 to 30-May-2026 (Meetings)"
    ),
}

_RANGE_RE = re.compile(r"\s+to\s+", re.I)


def _format_date(val: Any) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%m/%d/%Y")
    return str(val).strip()


def _add_months(start: datetime, months: int) -> datetime:
    """Add calendar months; end date is last day of term (slide style)."""
    y = start.year + (start.month - 1 + months) // 12
    m = (start.month - 1 + months) % 12 + 1
    d = min(start.day, 28)
    try:
        return datetime(y, m, d)
    except ValueError:
        return datetime(y, m, 28)


def _parse_single_date(text: str) -> datetime | None:
    text = text.strip()
    for fmt in ("%m/%d/%Y", "%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def derive_subscription_dates(
    sub_start: Any,
    sub_term_months: Any,
    *,
    existing_sub_dates: Any = None,
) -> str:
    if existing_sub_dates not in (None, ""):
        return str(existing_sub_dates).strip()

    if sub_start in (None, ""):
        return ""

    if isinstance(sub_start, datetime):
        start_str = _format_date(sub_start)
        if sub_term_months not in (None, ""):
            try:
                months = int(float(str(sub_term_months).strip()))
                end = _add_months(sub_start, months)
                return f"{start_str} to {end.strftime('%m/%d/%Y')}"
            except (ValueError, TypeError):
                pass
        return start_str

    start_text = str(sub_start).strip()
    if _RANGE_RE.search(start_text):
        return start_text

    if sub_term_months not in (None, ""):
        try:
            months = int(float(str(sub_term_months).strip()))
        except (ValueError, TypeError):
            months = 0
        if months > 0:
            first_part = start_text.split(",")[0].strip()
            parsed = _parse_single_date(first_part)
            if parsed:
                end = _add_months(parsed, months)
                return f"{parsed.strftime('%m/%d/%Y')} to {end.strftime('%m/%d/%Y')}"

    return start_text


def _insert_column_after(
    ws, col_index: dict[str, int], after_header: str
) -> tuple[int, dict[str, int]]:
    existing = _col_for_field(col_index, SUBSCRIPTION_DATES_COL)
    if existing:
        return existing, col_index

    after_col = _col_for_field(col_index, after_header) or max(col_index.values(), default=2)
    insert_at = after_col + 1

    ws.insert_cols(insert_at)
    section = ws.cell(1, after_col).value or "Subscription Review"
    ws.cell(1, insert_at, value=section if section else "Subscription Review")
    ws.cell(1, insert_at).fill = SECTION_FILL
    ws.cell(2, insert_at, value=SUBSCRIPTION_DATES_COL)
    ws.cell(2, insert_at).fill = HEADER_FILL
    ws.cell(2, insert_at).font = HEADER_FONT
    ws.cell(2, insert_at).alignment = Alignment(wrap_text=True, vertical="top")

    updated: dict[str, int] = {}
    for name, col in col_index.items():
        updated[name] = col + 1 if col >= insert_at else col
    updated[SUBSCRIPTION_DATES_COL] = insert_at
    return insert_at, updated


def sync_subscription_dates(path: Path) -> dict[str, Any]:
    path = Path(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_index = _header_map(ws)

    dates_col = _col_for_field(col_index, SUBSCRIPTION_DATES_COL)
    if not dates_col:
        after = "Sub term (months)" if _col_for_field(col_index, "Sub term (months)") else "Platforms"
        dates_col, col_index = _insert_column_after(ws, col_index, after)

    filled = 0
    for row in range(3, ws.max_row + 1):
        row_data = {}
        for name, col in col_index.items():
            row_data[name] = ws.cell(row, col).value
        org = str(row_data.get("Customer org id") or "").strip().lower()

        if org in VERIFIED_SUBSCRIPTION_DATES:
            value = VERIFIED_SUBSCRIPTION_DATES[org]
        else:
            value = derive_subscription_dates(
                row_data.get("Sub start date (MM/DD/YYYY)"),
                row_data.get("Sub term (months)"),
                existing_sub_dates=row_data.get(SUBSCRIPTION_DATES_COL),
            )

        if value:
            ws.cell(row, dates_col, value=value)
            filled += 1

    wb.save(path)
    return {
        "path": str(path),
        "column": SUBSCRIPTION_DATES_COL,
        "column_index": dates_col,
        "rows_filled": filled,
    }

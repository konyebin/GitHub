"""Merge PDF/screenshot/install-base sources into Check Back workbook."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from src.ingest.install_base_mapper import map_row
from src.ingest.install_base_mapper import load_config as load_map_config
from src.workbook.check_back_template import append_row, flatten_merged_fields, load_template
from src.workbook.note_parser import compute_active_pct, parse_notes_and_cells


PROV_ENT_LICENSE_COL = "Provisioned/Entitled Lic Calling"


def _license_cells(row: dict[str, Any]) -> tuple[Any, Any]:
    """Return (provisioned-ish, active) sources for pct — prefers merged column."""
    merged = row.get(PROV_ENT_LICENSE_COL)
    if merged not in (None, ""):
        return merged, row.get("Active Lic Calling")
    return row.get("Providioned Lic Calling"), row.get("Active Lic Calling")


def enrich_row_from_notes(row: dict[str, Any]) -> dict[str, Any]:
    """Add PowerPoint-aligned structured fields from note columns when empty."""
    analytics = str(row.get("Notes from Calling Analytics") or "")
    features = str(row.get("Notes from provisioned features") or "")
    addons = str(row.get("Add-ons included or not") or "")
    parsed = parse_notes_and_cells(analytics, features, addons, row)
    pct = compute_active_pct(*_license_cells(row))
    if pct:
        parsed["Active % of provisioned"] = pct
    for k, v in parsed.items():
        if k not in row or row[k] in (None, ""):
            row[k] = v
    return row


def clear_data_rows(ws, keep_through_row: int = 2) -> None:
    """Remove existing data rows; keep section + header rows."""
    if ws.max_row > keep_through_row:
        ws.delete_rows(keep_through_row + 1, ws.max_row - keep_through_row)


def rows_from_customers_json(staging_dir: str | Path) -> list[tuple[dict[str, Any], set[str]]]:
    path = Path(staging_dir) / "customers.json"
    if not path.exists():
        merged = Path(staging_dir) / "merged_extracted.json"
        if merged.exists():
            data = json.loads(merged.read_text(encoding="utf-8"))
            return [flatten_merged_fields(data.get("merged_fields", {}))]
        return []

    records = json.loads(path.read_text(encoding="utf-8"))
    return [flatten_merged_fields(r.get("merged_fields", {})) for r in records]


def build_workbook(
    template_path: str | Path,
    output_path: str | Path,
    staging_dir: str | Path | None = None,
    install_base_path: str | Path | None = None,
    account_filter: str | None = None,
    ib_limit: int | None = None,
    mode: str = "fresh",
) -> Path:
    """
    Build Check Back xlsx from ingested sources.

    mode=fresh (default): template headers only + rows from PDFs/screenshots (and optional IB).
    mode=append: keep existing template rows and append source rows.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    staging_dir = Path(staging_dir) if staging_dir else None

    wb, col_index = load_template(template_path)
    ws = wb.active

    if mode == "fresh":
        clear_data_rows(ws)

    if staging_dir:
        for row, low in rows_from_customers_json(staging_dir):
            if account_filter:
                row["Opportunity Name"] = account_filter
            if account_filter:
                name = str(row.get("Opportunity Name", "")).lower()
                if account_filter.lower() not in name and name:
                    continue
            append_row(ws, col_index, enrich_row_from_notes(row), low)

    if install_base_path:
        import openpyxl

        config = load_map_config()
        ib = openpyxl.load_workbook(install_base_path, read_only=True, data_only=True)
        ws_ib = ib.active
        headers = [c.value for c in next(ws_ib.iter_rows(min_row=1, max_row=1))]
        count = 0
        for row in ws_ib.iter_rows(min_row=2, values_only=True):
            if ib_limit is not None and count >= ib_limit:
                break
            ib_row = dict(zip(headers, row))
            if not any(v not in (None, "") for v in ib_row.values()):
                continue
            cb_row, _ = map_row(ib_row, config)
            if account_filter:
                name = str(cb_row.get("Opportunity Name", "")).lower()
                if account_filter.lower() not in name and name:
                    continue
            append_row(ws, col_index, enrich_row_from_notes(cb_row))
            count += 1
        ib.close()

    if ws.max_row <= 2 and account_filter:
        append_row(ws, col_index, {"Opportunity Name": account_filter})

    wb.save(output_path)
    wb.close()
    return output_path


def validate_workbook(path: str | Path) -> list[dict[str, str]]:
    """Return validation issues for Check Back rows."""
    import openpyxl

    issues: list[dict[str, str]] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[2]]
    col = {h: i for i, h in enumerate(headers) if h}

    def num_from(s: Any) -> float:
        if not s:
            return 0
        import re

        m = re.findall(r"[\d,]+", str(s))
        vals = [float(x.replace(",", "")) for x in m if x.replace(",", "").strip()]
        return max(vals, default=0) if vals else 0

    def license_totals(row_vals: tuple[Any, ...]) -> tuple[float, float]:
        import re as _re

        if PROV_ENT_LICENSE_COL in col:
            cell = str(row_vals[col[PROV_ENT_LICENSE_COL]] or "")
            prov_nums = _re.findall(
                r"\b(?:PL|WS|Professional|Workspace)\s*:?\s*(\d[\d,]*)\s*/\s*\d[\d,]*",
                cell,
                _re.I,
            )
            ent_nums = _re.findall(
                r"\b(?:PL|WS|Professional|Workspace)\s*:?\s*\d[\d,]*\s*/\s*(\d[\d,]*)",
                cell,
                _re.I,
            )
            if prov_nums or ent_nums:
                prov = sum(float(x.replace(",", "")) for x in prov_nums)
                ent = sum(float(x.replace(",", "")) for x in ent_nums)
                return ent, prov
        ent = num_from(
            row_vals[col["Entitled Lic Calling"]] if "Entitled Lic Calling" in col else ""
        )
        prov = num_from(
            row_vals[col["Providioned Lic Calling"]] if "Providioned Lic Calling" in col else ""
        )
        return ent, prov

    for r_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not any(row):
            continue
        name = row[col.get("Opportunity Name", 0)] if "Opportunity Name" in col else ""
        ent, prov = license_totals(row)
        act = num_from(row[col["Active Lic Calling"]] if "Active Lic Calling" in col else "")
        if act > prov > 0 and prov > 0:
            issues.append(
                {"row": str(r_idx), "account": str(name), "issue": "Active exceeds Provisioned"}
            )
        if ent > 0 and act > 0 and act > ent * 1.1:
            issues.append(
                {
                    "row": str(r_idx),
                    "account": str(name),
                    "issue": "Active much higher than Entitled — verify",
                }
            )
        if "Customer org id" in col and not row[col["Customer org id"]] and name:
            issues.append(
                {"row": str(r_idx), "account": str(name), "issue": "Missing Customer org id"}
            )

    wb.close()
    return issues

"""Normalize Check Back workbook column values per config/column_standards.yaml."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

from .sub_term import LEGACY_SUB_START_COL, SUB_TERM_COL, sub_term_from_text

PROV_ENT_LICENSE_COL = "Provisioned/Entitled Lic Calling"

RED_FONT = Font(color="FF0000")
UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
    re.I,
)
SFDC_URL_RE = re.compile(
    r"^https://ciscocollaboration\.lightning\.force\.com/",
    re.I,
)
SUB_RE = re.compile(r"Sub\s*(\d+)", re.I)
DATE_RANGE_RE = re.compile(
    r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{2}-\d{2}|\d{1,2}-[A-Za-z]{3}-\d{4})"
    r"\s*(?:to|-|–)\s*"
    r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{2}-\d{2}|\d{1,2}-[A-Za-z]{3}-\d{4})",
    re.I,
)
PL_WS_EXPLICIT = re.compile(
    r"PL\s*:\s*([\d,]+)\s*,?\s*WS\s*:\s*([\d,]+)",
    re.I,
)
PL_ONLY = re.compile(r"\bPL\s*:\s*([\d,]+)", re.I)
WS_ONLY = re.compile(r"\bWS\s*:\s*([\d,]+)", re.I)
PL_WS_WORKSPACE = re.compile(
    r"^([\d,]+)\s+workspace\s+([\d,]+)$",
    re.I,
)
PROF_WS_PAIR_RE = re.compile(
    r"\b(?:Professional|Workspace|PL|WS)\s*:?\s*\d[\d,]*\s*/\s*\d[\d,]*",
    re.I,
)
ACTIVE_RE = re.compile(r"^([\d,]+)\s*(?:active)?$", re.I)
CCEP_DATE_RE = re.compile(r"^([YyNn])\s*\(([^)]+)\)\s*$")

SL2_CANONICAL = {
    "us commercial": "US Commercial",
    "us ps market segment": "US PS Market Segment",
    "emea__uki": "EMEA__UKI",
    "sled-west area": "SLED-West Area",
}

MIGRATING_FROM_MAP = {
    "cucm": "CUCM",
    "ucm, uccx": "UCM, UCCX",
    "consolidating": "Consolidating",
    "n/a": "N/A",
    "avaya": "Avaya",
    "hcs": "HCS",
}

MIGRATING_TO_MAP = {
    "wxc": "WxC",
    "wxcc": "WxCC",
    "wxc, wxcc": "WxC, WxCC",
    "wxc wxcc": "WxC, WxCC",
    "wxc-wxcc": "WxC, WxCC",
    "wxc di": "WxC DI",
    "webex suite": "Webex Suite",
}


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _set_cell(ws, row: int, col: int, value: str | int | float, red: bool = False) -> None:
    cell = ws.cell(row, col)
    cell.value = value
    if red:
        cell.font = RED_FONT


def _set_url_cell(ws, row: int, col: int, url: str, red: bool = False) -> None:
    cell = ws.cell(row, col)
    cell.value = url
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url)
    if red:
        cell.font = RED_FONT
    else:
        cell.font = Font(color="0563C1", underline="single")


def _salesforce_from_cell(cell, raw: str) -> tuple[str, bool]:
    hl = cell.hyperlink
    if hl and hl.target:
        target = str(hl.target).strip()
        if SFDC_URL_RE.match(target) or "lightning.force.com" in target:
            return target, False
    return _salesforce_url_or_red(raw)


def _parse_iso_date(s: str) -> str | None:
    s = s.strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def _format_currency(val: Any) -> tuple[str, bool]:
    s = _cell_str(val).replace("$", "").replace(",", "").strip()
    if not s:
        return "", False
    try:
        n = int(float(s))
    except ValueError:
        return _cell_str(val), True
    return f"${n:,}", False


def _normalize_sub_numbers(text: str) -> str:
    if not text:
        return text
    parts = re.split(r"[\n,]+", text)
    out = []
    for part in parts:
        p = part.strip()
        if not p:
            continue
        p = re.sub(r"Sub(\d+)", r"Sub\1", p, flags=re.I)
        p = re.sub(r"sub(\d+)", lambda m: f"Sub{m.group(1)}", p, flags=re.I)
        out.append(p)
    return ", ".join(out)


def _uuid_or_red(s: str) -> tuple[str, bool]:
    s = s.strip().lower()
    if not s:
        return "", False
    if UUID_RE.match(s):
        return s, False
    return _cell_str(s), True


def _salesforce_url_or_red(s: str) -> tuple[str, bool]:
    s = _cell_str(s)
    if not s:
        return "", False
    if SFDC_URL_RE.match(s):
        return s, False
    placeholders = {"salesforce link", "salesfroce link", "s&c", "s and c"}
    if s.lower() in placeholders or not s.startswith("http"):
        return s, True
    return s, not SFDC_URL_RE.match(s)


def _gyr(s: str) -> tuple[str, bool]:
    s = _cell_str(s)
    if not s:
        return "", False
    letter = s.upper()[:1]
    if letter in ("G", "Y", "R"):
        return letter, False
    return s, True


def _yn_only(s: str) -> tuple[str, bool]:
    sl = _cell_str(s).upper()
    if not sl:
        return "", False
    if sl in ("Y", "N"):
        return sl, False
    if sl in ("YES", "NO"):
        return "Y" if sl == "YES" else "N", False
    return _cell_str(s), True


def _ccep_yn(s: str) -> tuple[str, str | None, bool]:
    raw = _cell_str(s)
    if not raw:
        return "", None, False
    m = CCEP_DATE_RE.match(raw)
    if m:
        letter = m.group(1).upper()
        note = f"[CCEP trial date: {m.group(2).strip()}]"
        return letter, note, False
    yn, red = _yn_only(raw)
    return yn, None, red


def _csm_model(s: str) -> tuple[str, bool]:
    raw = _cell_str(s)
    if not raw:
        return "", False
    key = raw.lower()
    if key in ("scale", "tac support", "high touch", "none"):
        mapping = {
            "scale": "Scale",
            "tac support": "TAC Support",
            "high touch": "High Touch",
            "none": "None",
        }
        return mapping[key], False
    if raw == "TAC Supported":
        return "TAC Support", False
    return raw, False


def _migrating_from(s: str) -> tuple[str, bool]:
    raw = _cell_str(s)
    if not raw:
        return "", False
    key = raw.lower()
    if key in MIGRATING_FROM_MAP:
        return MIGRATING_FROM_MAP[key], False
    return raw, False


def _migrating_to(s: str) -> tuple[str, bool]:
    raw = _cell_str(s)
    if not raw:
        return "", False
    key = re.sub(r"\s+", " ", raw.lower())
    if key in MIGRATING_TO_MAP:
        return MIGRATING_TO_MAP[key], False
    return raw, False


def _competitor(s: str) -> tuple[str, bool]:
    raw = _cell_str(s)
    if not raw:
        return "", False
    if raw.lower() == "none":
        return "None", False
    parts = re.split(r"[;,]", raw)
    cleaned = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if p.lower() == "microsoft":
            cleaned.append("Microsoft")
        elif p.lower() == "zoom":
            cleaned.append("Zoom")
        elif p.lower() == "8x8":
            cleaned.append("8x8")
        else:
            cleaned.append(p)
    return ";".join(cleaned), False


def _sl2(s: str) -> tuple[str, bool]:
    raw = _cell_str(s)
    if not raw:
        return "", False
    key = raw.lower()
    if key in SL2_CANONICAL:
        return SL2_CANONICAL[key], False
    return raw, False


def _fmt_count(n: int) -> str:
    return f"{int(n):,}"


def _parse_pl_ws_single(raw: str) -> tuple[int | None, int | None]:
    text = _cell_str(raw)
    if not text or text.lower() in ("none", "no"):
        return None, None
    pl: int | None = None
    ws: int | None = None
    m = PL_WS_EXPLICIT.search(text)
    if m:
        return int(m.group(1).replace(",", "")), int(m.group(2).replace(",", ""))
    m = PL_WS_WORKSPACE.match(text)
    if m:
        return int(m.group(1).replace(",", "")), int(m.group(2).replace(",", ""))
    m = PL_ONLY.search(text)
    if m:
        pl = int(m.group(1).replace(",", ""))
    m = WS_ONLY.search(text)
    if m:
        ws = int(m.group(1).replace(",", ""))
    if pl is not None or ws is not None:
        return pl, ws
    if re.fullmatch(r"[\d,]+", text.replace(",", "")):
        return int(text.replace(",", "")), None
    return None, None


def _has_prof_ws_pair_format(raw: str) -> bool:
    return bool(PROF_WS_PAIR_RE.search(_cell_str(raw)))


def _normalize_prof_ws_display(text: str) -> str:
    s = _cell_str(text)
    if not s:
        return ""

    def _repl_pl(m: re.Match[str]) -> str:
        return f"Professional {_fmt_count(int(m.group(1).replace(',', '')))}/{_fmt_count(int(m.group(2).replace(',', '')))}"

    def _repl_ws(m: re.Match[str]) -> str:
        return f"Workspace {_fmt_count(int(m.group(1).replace(',', '')))}/{_fmt_count(int(m.group(2).replace(',', '')))}"

    s = re.sub(
        r"\bPL\s*:?\s*([\d,]+)\s*/\s*([\d,]+)",
        _repl_pl,
        s,
        flags=re.I,
    )
    s = re.sub(
        r"\bWS\s*:?\s*([\d,]+)\s*/\s*([\d,]+)",
        _repl_ws,
        s,
        flags=re.I,
    )
    s = re.sub(
        r"\bProfessional\s+([\d,]+)\s*/\s*([\d,]+)",
        _repl_pl,
        s,
        flags=re.I,
    )
    s = re.sub(
        r"\bWorkspace\s+([\d,]+)\s*/\s*([\d,]+)",
        _repl_ws,
        s,
        flags=re.I,
    )
    return re.sub(r"\s*;\s*", "; ", s.replace(", Workspace", "; Workspace"))


def _format_entitled_merged(ent_raw: str, prov_raw: str) -> str:
    entitled = _cell_str(ent_raw)
    provisioned = _cell_str(prov_raw)
    if _has_prof_ws_pair_format(entitled):
        return _normalize_prof_ws_display(entitled)
    if _has_prof_ws_pair_format(provisioned) and not entitled:
        return _normalize_prof_ws_display(provisioned)

    ent_pl, ent_ws = _parse_pl_ws_single(entitled)
    prov_pl, prov_ws = _parse_pl_ws_single(provisioned)
    parts: list[str] = []
    if ent_pl is not None or prov_pl is not None:
        parts.append(
            f"Professional {_fmt_count(prov_pl or 0)}/{_fmt_count(ent_pl or 0)}"
        )
    if ent_ws is not None or prov_ws is not None:
        parts.append(
            f"Workspace {_fmt_count(prov_ws or 0)}/{_fmt_count(ent_ws or 0)}"
        )
    if parts:
        return "; ".join(parts)
    return entitled


def _format_provisioned_column(prov_raw: str) -> str:
    raw = _cell_str(prov_raw)
    if not raw or raw.lower() in ("none", "no"):
        return ""
    if _has_prof_ws_pair_format(raw):
        return _normalize_prof_ws_display(raw)
    prov_pl, prov_ws = _parse_pl_ws_single(raw)
    parts: list[str] = []
    if prov_pl is not None:
        parts.append(f"Professional {_fmt_count(prov_pl)}")
    if prov_ws is not None:
        parts.append(f"Workspace {_fmt_count(prov_ws)}")
    if parts:
        return "; ".join(parts)
    return raw


def _pl_ws_license(s: str) -> tuple[str, bool]:
    raw = _cell_str(s)
    if not raw:
        return "", False
    if raw.lower() == "no":
        return "No", False

    m = PL_WS_EXPLICIT.search(raw)
    if m:
        pl = m.group(1).replace(",", "")
        ws = m.group(2).replace(",", "")
        return f"PL: {int(pl):,}, WS: {int(ws):,}", False

    m = PL_WS_WORKSPACE.match(raw)
    if m:
        pl = int(m.group(1).replace(",", ""))
        ws = int(m.group(2).replace(",", ""))
        return f"PL: {pl:,}, WS: {ws:,}", False

    if re.fullmatch(r"[\d,]+$", raw.replace(",", "")):
        n = int(raw.replace(",", ""))
        return f"PL: {n:,}", False

    return raw, True


def _active_license(s: str) -> tuple[str, bool]:
    raw = _cell_str(s)
    if not raw:
        return "", False
    m = ACTIVE_RE.match(raw.replace(",", ""))
    if m:
        return _fmt_count(int(m.group(1).replace(",", ""))), False
    if raw.lower() == "no":
        return "0", False
    if re.fullmatch(r"[\d,]+", raw.replace(",", "")):
        return _fmt_count(int(raw.replace(",", ""))), False
    return raw, True


def _date_iso(val: Any) -> tuple[str, bool]:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d"), False
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d"), False
    s = _cell_str(val)
    if not s:
        return "", False
    if "00:00:00" in s:
        s = s.split(" ")[0]
    iso = _parse_iso_date(s)
    if iso:
        return iso, False
    return s, True


def _date_range_or_red(s: str) -> tuple[str, bool]:
    raw = _cell_str(s)
    if not raw:
        return "", False
    if " to " in raw.lower() or re.search(r"\d\s*-\s*\d", raw):
        m = DATE_RANGE_RE.search(raw)
        if m:
            d1 = _parse_iso_date(m.group(1))
            d2 = _parse_iso_date(m.group(2))
            if d1 and d2:
                return f"{d1} to {d2}", False
        return raw, True
    lone = _parse_iso_date(raw.split(" ")[0].replace("00:00:00", "").strip())
    if lone:
        return raw, True
    return raw, True


def detect_header_row(ws, max_scan: int = 5) -> int:
    for r in range(1, max_scan + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and str(v).strip() == "Opportunity Name":
                return r
    return 2


def standardize_workbook(
    input_path: str | Path,
    output_path: str | Path | None = None,
    as_of: date | None = None,
) -> dict[str, Any]:
    input_path = Path(input_path)
    output_path = Path(output_path) if output_path else input_path.with_name(
        input_path.stem + "_standardized" + input_path.suffix
    )

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    header_row = detect_header_row(ws)

    headers: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v and str(v).strip():
            headers[c] = str(v).strip()

    col_by_name = {h: c for c, h in headers.items()}
    notes_col = col_by_name.get("Notes from Calling Analytics")
    entitled_col = col_by_name.get("Entitled Lic Calling")
    provisioned_col = col_by_name.get("Providioned Lic Calling")
    prov_ent_col = col_by_name.get(PROV_ENT_LICENSE_COL)
    term_as_of = as_of or date.today()

    # Rename legacy Sub start column → Sub Term
    for c, h in list(headers.items()):
        if h == LEGACY_SUB_START_COL:
            ws.cell(header_row, c).value = SUB_TERM_COL
            headers[c] = SUB_TERM_COL
            col_by_name[SUB_TERM_COL] = c
            del col_by_name[LEGACY_SUB_START_COL]
            break

    stats: dict[str, Any] = {
        "input": str(input_path),
        "output": str(output_path),
        "header_row": header_row,
        "rows_processed": 0,
        "cells_changed": 0,
        "cells_red": 0,
        "red_by_column": {},
        "as_of": term_as_of.isoformat(),
    }

    for r in range(header_row + 1, ws.max_row + 1):
        if not any(ws.cell(r, c).value for c in headers):
            continue
        stats["rows_processed"] += 1

        for c, name in headers.items():
            cell = ws.cell(r, c)
            val = cell.value
            raw = _cell_str(val)
            new_val: str | int | float = raw
            red = False
            url_cell = False

            if name == "Opportunity Name" or name == "Partner" or name == "CSM name":
                new_val = raw.strip()
            elif name == "Opportunity (linked)":
                new_val, red = _salesforce_from_cell(cell, raw)
                url_cell = bool(new_val) and not red and str(new_val).startswith("http")
            elif name == "Customer org id":
                new_val, red = _uuid_or_red(raw)
            elif name == "Sub #":
                new_val = _normalize_sub_numbers(raw)
            elif name == "TCV $":
                new_val, red = _format_currency(val)
            elif name == "Closed on (MM/YY)":
                new_val, red = _date_iso(val)
            elif name in (SUB_TERM_COL, LEGACY_SUB_START_COL):
                new_val, red = sub_term_from_text(raw, term_as_of)
            elif name == " (G/Y/R)" or name == "(G/Y/R)":
                new_val, red = _gyr(raw)
            elif name == "CCEP trial (Y/N)":
                yn, note, red = _ccep_yn(raw)
                new_val = yn
                if note and notes_col:
                    existing = _cell_str(ws.cell(r, notes_col).value)
                    ws.cell(r, notes_col).value = (existing + " " + note).strip()
            elif name == "Calling Setup Assist included (Y/N)":
                new_val, red = _yn_only(raw)
            elif name == "CSM Engagement Model (linked)":
                new_val, _ = _csm_model(raw)
            elif name == "Migrating from":
                new_val, _ = _migrating_from(raw)
            elif name == "Migrating to":
                new_val, _ = _migrating_to(raw)
            elif name == "Competitor":
                new_val, _ = _competitor(raw)
            elif name == "SL2":
                new_val, _ = _sl2(raw)
            elif name == PROV_ENT_LICENSE_COL:
                if _has_prof_ws_pair_format(raw):
                    new_val = _normalize_prof_ws_display(raw)
                else:
                    new_val, red = _pl_ws_license(raw)
            elif name == "Entitled Lic Calling" or name == "Providioned Lic Calling":
                new_val, red = _pl_ws_license(raw)
            elif name == "Active Lic Calling":
                new_val, red = _active_license(raw)
            elif name == "TAC/BEMS":
                new_val, red = _yn_only(raw)
            else:
                continue

            if new_val != raw or red or url_cell:
                if url_cell:
                    _set_url_cell(ws, r, c, str(new_val), red=red)
                else:
                    _set_cell(ws, r, c, new_val, red=red)
                stats["cells_changed"] += 1
                if red:
                    stats["cells_red"] += 1
                    stats["red_by_column"][name] = stats["red_by_column"].get(name, 0) + 1

        ent_raw = _cell_str(ws.cell(r, entitled_col).value) if entitled_col else ""
        prov_raw = _cell_str(ws.cell(r, provisioned_col).value) if provisioned_col else ""
        merged_ent = _format_entitled_merged(ent_raw, prov_raw)

        if prov_ent_col:
            pe_raw = _cell_str(ws.cell(r, prov_ent_col).value)
            if pe_raw and _has_prof_ws_pair_format(pe_raw):
                normalized = _normalize_prof_ws_display(pe_raw)
                if normalized != pe_raw:
                    _set_cell(ws, r, prov_ent_col, normalized)
                    stats["cells_changed"] += 1
            elif not pe_raw and merged_ent:
                _set_cell(ws, r, prov_ent_col, merged_ent)
                stats["cells_changed"] += 1
            elif pe_raw and not _has_prof_ws_pair_format(pe_raw) and (ent_raw or prov_raw):
                if merged_ent and merged_ent != pe_raw:
                    _set_cell(ws, r, prov_ent_col, merged_ent)
                    stats["cells_changed"] += 1
        elif entitled_col or provisioned_col:
            merged_prov = _format_provisioned_column(prov_raw)
            if entitled_col and merged_ent and merged_ent != ent_raw:
                _set_cell(ws, r, entitled_col, merged_ent)
                stats["cells_changed"] += 1
            if provisioned_col and merged_prov != prov_raw:
                _set_cell(ws, r, provisioned_col, merged_prov or "")
                stats["cells_changed"] += 1

    wb.save(output_path)
    return stats

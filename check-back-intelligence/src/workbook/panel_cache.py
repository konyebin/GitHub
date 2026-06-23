"""Export / import Subscription Review, Provisioning, and Feature Use panel columns."""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .slide_layout import FEATURE_FIELDS, PROVISIONING_FIELDS, SUBSCRIPTION_FIELDS

SECTION_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FILL = PatternFill(start_color="0D1526", end_color="0D1526", fill_type="solid")
HEADER_FONT = Font(bold=True, color="94A3B8", size=10)

MATCH_KEYS = ("Opportunity Name", "Customer org id")

# Columns shown in the three BIA panels (plus keys to match rows on import).
PANEL_COLUMNS: list[tuple[str, str]] = [
    ("Opportunity Name", "Keys"),
    ("Customer org id", "Keys"),
    # Subscription Review
    ("Sub #", "Subscription Review"),
    ("TCV $", "Subscription Review"),
    ("AAR $", "Subscription Review"),
    ("Partner", "Subscription Review"),
    ("SL2", "Subscription Review"),
    ("Collab AE/SE", "Subscription Review"),
    ("Platforms", "Subscription Review"),
    ("Service lines", "Subscription Review"),
    ("Sub term (months)", "Subscription Review"),
    ("Subscription dates", "Subscription Review"),
    ("Trend active users 90d", "Subscription Review"),
    ("Trend call volume 90d", "Subscription Review"),
    ("Salesforce URL", "Subscription Review"),
    *[(name, "Subscription Review") for name in SUBSCRIPTION_FIELDS if name not in {
        "AAR $", "Collab AE/SE", "Platforms", "Service lines", "Sub term (months)",
        "Subscription dates", "Trend active users 90d", "Trend call volume 90d", "Salesforce URL",
    }],
    # Provisioning & Usage
    ("Provisioned/Entitled Lic Calling", "Provisioning & Usage"),
    ("Entitled Lic Calling", "Provisioning & Usage"),
    ("Providioned Lic Calling", "Provisioning & Usage"),
    ("Active Lic Calling", "Provisioning & Usage"),
    ("Notes from Calling Analytics", "Provisioning & Usage"),
    ("Notes from provisioned features", "Provisioning & Usage"),
    *[(name, "Provisioning & Usage") for name in PROVISIONING_FIELDS],
    ("Total calls", "Provisioning & Usage"),
    ("Answered calls %", "Provisioning & Usage"),
    ("Calls busiest hour", "Provisioning & Usage"),
    # Feature Use & Add-ons
    ("Add-ons included or not", "Feature Use & Add-ons"),
    *[(name, "Feature Use & Add-ons") for name in FEATURE_FIELDS],
]

# De-dupe while preserving order
_seen: set[str] = set()
PANEL_COLUMN_LIST: list[tuple[str, str]] = []
for name, section in PANEL_COLUMNS:
    if name in _seen:
        continue
    _seen.add(name)
    PANEL_COLUMN_LIST.append((name, section))

PANEL_FIELD_NAMES = [name for name, _ in PANEL_COLUMN_LIST if name not in MATCH_KEYS]


def default_panel_cache_path(workbook_path: str | Path) -> Path:
    p = Path(workbook_path)
    return p.with_name(f"{p.stem}_panels.xlsx")


def _header_map(ws, header_row: int = 2) -> dict[str, int]:
    m: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if not v:
            continue
        key = str(v).strip()
        if key not in m:
            m[key] = col
    return m


def _detect_header_row(ws) -> int:
    for row in (2, 1, 3):
        vals = _header_map(ws, row)
        if "Opportunity Name" in vals:
            return row
    return 2


def _norm(s: Any) -> str:
    return str(s or "").strip().lower()


def _row_key(row: dict[str, Any]) -> tuple[str, str]:
    return (_norm(row.get("Opportunity Name")), _norm(row.get("Customer org id")))


def _read_rows(ws, header_row: int) -> list[dict[str, Any]]:
    col_index = _header_map(ws, header_row)
    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {h: ws.cell(r, col_index[h]).value for h in col_index}
        if not any(v not in (None, "") for v in row.values()):
            continue
        rows.append(row)
    return rows


def _write_panel_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Panel Data"
    seen_sections: set[str] = set()
    for col, (name, section) in enumerate(PANEL_COLUMN_LIST, start=1):
        ws.cell(1, col, value=section if section not in seen_sections else "")
        seen_sections.add(section)
        ws.cell(2, col, value=name)
        ws.cell(1, col).fill = SECTION_FILL
        ws.cell(2, col).fill = HEADER_FILL
        ws.cell(2, col).font = HEADER_FONT
        ws.cell(2, col).alignment = Alignment(wrap_text=True, vertical="top")
    for r_idx, src in enumerate(rows, start=3):
        for col, (name, _) in enumerate(PANEL_COLUMN_LIST, start=1):
            val = src.get(name)
            if val not in (None, ""):
                ws.cell(r_idx, col, value=val)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _ensure_columns(ws, col_index: dict[str, int], header_row: int) -> dict[str, int]:
    start = max(col_index.values(), default=header_row) + 1
    seen_sections: set[str] = set()
    panel_names = {name for name, _ in PANEL_COLUMN_LIST}
    for name, section in PANEL_COLUMN_LIST:
        if name in col_index:
            continue
        ws.cell(1, start, value=section if section not in seen_sections else "")
        seen_sections.add(section)
        ws.cell(2, start, value=name)
        ws.cell(1, start).fill = SECTION_FILL
        ws.cell(2, start).fill = HEADER_FILL
        ws.cell(2, start).font = HEADER_FONT
        col_index[name] = start
        start += 1
    # silence unused warning
    _ = panel_names
    return col_index


def export_panels(
    workbook_path: str | Path,
    output_path: str | Path | None = None,
    *,
    backup: bool = False,
) -> dict[str, Any]:
    """Copy panel-column values from the main Check Back workbook into a sidecar xlsx."""
    workbook_path = Path(workbook_path)
    output_path = Path(output_path) if output_path else default_panel_cache_path(workbook_path)
    if backup and output_path.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(output_path, output_path.with_name(output_path.stem + f".backup_{ts}" + output_path.suffix))

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb.active
    header_row = _detect_header_row(ws)
    col_index = _header_map(ws, header_row)
    src_rows = _read_rows(ws, header_row)

    out_rows: list[dict[str, Any]] = []
    for src in src_rows:
        out: dict[str, Any] = {}
        for name, _ in PANEL_COLUMN_LIST:
            if name in col_index:
                out[name] = src.get(name)
            else:
                out[name] = None
        if any(out.get(k) not in (None, "") for k in MATCH_KEYS):
            out_rows.append(out)

    _write_panel_workbook(output_path, out_rows)
    return {
        "workbook": str(workbook_path),
        "output": str(output_path),
        "rows": len(out_rows),
        "columns": len(PANEL_COLUMN_LIST),
    }


def import_panels(
    workbook_path: str | Path,
    panel_path: str | Path | None = None,
    *,
    backup: bool = True,
) -> dict[str, Any]:
    """Merge panel sidecar xlsx values back into the main Check Back workbook."""
    workbook_path = Path(workbook_path)
    panel_path = Path(panel_path) if panel_path else default_panel_cache_path(workbook_path)
    if not panel_path.exists():
        raise FileNotFoundError(f"Panel cache not found: {panel_path}")

    if backup:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(workbook_path, workbook_path.with_name(workbook_path.stem + f".panel_backup_{ts}" + workbook_path.suffix))

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active
    header_row = _detect_header_row(ws)
    col_index = _header_map(ws, header_row)
    col_index = _ensure_columns(ws, col_index, header_row)

    panel_wb = openpyxl.load_workbook(panel_path, data_only=True)
    panel_ws = panel_wb.active
    panel_header_row = _detect_header_row(panel_ws)
    panel_rows = _read_rows(panel_ws, panel_header_row)
    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    by_name: dict[str, dict[str, Any]] = {}
    for row in panel_rows:
        key = _row_key(row)
        if key[0] or key[1]:
            by_key[key] = row
        name = _norm(row.get("Opportunity Name"))
        if name:
            by_name[name] = row

    matched = 0
    cells_written = 0
    for r in range(header_row + 1, ws.max_row + 1):
        main_row = {h: ws.cell(r, col_index[h]).value for h in col_index if h in col_index}
        key = _row_key(main_row)
        panel = by_key.get(key)
        if not panel and key[0]:
            panel = by_name.get(key[0])
        if not panel:
            continue
        matched += 1
        for field in PANEL_FIELD_NAMES:
            val = panel.get(field)
            if val in (None, ""):
                continue
            col = col_index.get(field)
            if not col:
                continue
            ws.cell(r, col, value=val)
            cells_written += 1

    wb.save(workbook_path)
    return {
        "workbook": str(workbook_path),
        "panel_cache": str(panel_path),
        "rows_matched": matched,
        "cells_written": cells_written,
    }

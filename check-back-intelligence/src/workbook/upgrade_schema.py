"""Add PowerPoint-aligned columns to Check Back workbook and backfill from notes."""

from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .note_parser import compute_active_pct, parse_notes_to_structured

# New columns appended after existing 24 (PowerPoint-aligned)
NEW_COLUMNS: list[tuple[str, str]] = [
    ("AAR $", "Subscription Review"),
    ("Collab AE/SE", "Subscription Review"),
    ("Service lines", "Subscription Review"),
    ("Sub term (months)", "Subscription Review"),
    ("Trend active users 90d", "Subscription Review"),
    ("Trend call volume 90d", "Subscription Review"),
    ("Lic Professional (used/entitled)", "Provisioning & Usage"),
    ("Lic Standard (used/entitled)", "Provisioning & Usage"),
    ("Lic Workspace (used/entitled)", "Provisioning & Usage"),
    ("Active % of provisioned", "Provisioning & Usage"),
    ("Total calls", "Provisioning & Usage"),
    ("External calls", "Provisioning & Usage"),
    ("Answered calls %", "Provisioning & Usage"),
    ("Calls busiest hour", "Provisioning & Usage"),
    ("Meetings usage", "Provisioning & Usage"),
    ("Messaging usage", "Provisioning & Usage"),
    ("Auto Attendant count", "Feature Use & Add-ons"),
    ("Hunt Groups count", "Feature Use & Add-ons"),
    ("Call Queues count", "Feature Use & Add-ons"),
    ("Connected-UC (Y/N)", "Feature Use & Add-ons"),
    ("Virtual Lines count", "Feature Use & Add-ons"),
    ("Recommended Actions", "Recommended Actions"),
    ("CSM / Account Team notes", "Recommended Actions"),
    ("Data gathered by", "Business Insight"),
    ("Data gathered date", "Business Insight"),
]

# Manual backfill for known pilot rows (from Business Insight slides)
PILOT_ROWS: dict[str, dict[str, Any]] = {
    "fb078f9f-4332-4ace-8b1f-6a46313aa4ea": {
        "AAR $": "261696",
        "Collab AE/SE": "Peter Caterinicchia",
        "Sub term (months)": "60",
        "Trend active users 90d": "UP +2.6%",
        "Trend call volume 90d": "UP +52%",
        "Lic Professional (used/entitled)": "4376/5193",
        "Lic Standard (used/entitled)": "0/0",
        "Lic Workspace (used/entitled)": "377/2164",
        "Active % of provisioned": "87%",
        "Total calls": "1.1M",
        "External calls": "35.8K out of 1.1M",
        "Answered calls %": "89.6%",
        "Meetings usage": "N/A",
        "Messaging usage": "1 active user; 1 message sent",
        "Auto Attendant count": "113",
        "Hunt Groups count": "9",
        "Call Queues count": "-",
        "Connected-UC (Y/N)": "Y",
        "Virtual Lines count": "96",
        "CSM / Account Team notes": (
            "Using Connected-UC; UCM cluster online but no users configured. "
            "Local Gateway for PSTN. Core calling usage strong. "
            "Adoption concern: SSO and Directory Connector not enabled."
        ),
        "Data gathered by": "Jeremy Abrams",
        "Data gathered date": "2026-05-19",
        " (G/Y/R)": "G",
    },
}

SECTION_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FILL = PatternFill(start_color="0D1526", end_color="0D1526", fill_type="solid")
HEADER_FONT = Font(bold=True, color="94A3B8", size=10)


def _existing_headers(ws) -> list[str]:
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(2, col).value
        if v:
            headers.append(str(v).strip())
    return headers


def upgrade_workbook(path: Path, backup: bool = True) -> dict[str, Any]:
    path = Path(path)
    if backup:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        bak = path.with_suffix(f".backup_{ts}.xlsx")
        shutil.copy2(path, bak)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = _existing_headers(ws)
    start_col = len(headers) + 1

    # Skip columns already present
    to_add = [(name, section) for name, section in NEW_COLUMNS if name not in headers]
    if not to_add:
        return {"added": 0, "message": "All PowerPoint columns already present"}

    col_index: dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        col_index[h] = i

    c = start_col
    section_cols: dict[str, int] = {}
    for name, section in to_add:
        ws.cell(1, c, value=section if section not in section_cols else "")
        ws.cell(2, c, value=name)
        ws.cell(1, c).fill = SECTION_FILL
        ws.cell(2, c).fill = HEADER_FILL
        ws.cell(2, c).font = HEADER_FONT
        ws.cell(2, c).alignment = Alignment(wrap_text=True, vertical="top")
        col_index[name] = c
        section_cols[section] = c
        c += 1

    filled = 0
    for row in range(3, ws.max_row + 1):
        row_data = {h: ws.cell(row, col_index[h]).value for h in col_index if h in col_index}
        org = str(row_data.get("Customer org id") or "").strip().lower()

        analytics = str(row_data.get("Notes from Calling Analytics") or "")
        features = str(row_data.get("Notes from provisioned features") or "")
        addons = str(row_data.get("Add-ons included or not") or "")

        parsed = parse_notes_to_structured(analytics, features, addons)
        pilot = PILOT_ROWS.get(org, {})

        active_pct = compute_active_pct(
            row_data.get("Providioned Lic Calling"),
            row_data.get("Active Lic Calling"),
        )
        if active_pct:
            parsed["Active % of provisioned"] = active_pct

        merged = {**parsed, **pilot}
        for field, val in merged.items():
            if field not in col_index or val in (None, ""):
                continue
            if ws.cell(row, col_index[field]).value in (None, ""):
                ws.cell(row, col_index[field], value=val)
                filled += 1

        # Split long notes into CSM box when empty
        csm_col = col_index.get("CSM / Account Team notes")
        if csm_col and not ws.cell(row, csm_col).value:
            tac = str(row_data.get("TAC/BEMS") or "")
            if tac and "adoption" in (analytics + features).lower():
                ws.cell(row, csm_col, value=tac[:500])

    wb.save(path)
    return {
        "path": str(path),
        "columns_added": len(to_add),
        "cells_filled": filled,
        "new_columns": [n for n, _ in to_add],
    }


if __name__ == "__main__":
    import sys

    target = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads/Check Back Initiative Customer Analysis.xlsx"
    result = upgrade_workbook(target)
    print(result)

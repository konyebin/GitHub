"""Rebuild PowerPoint-aligned columns from notes + verified slide data."""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .note_parser import compute_active_pct, parse_notes_to_structured
from .slide_layout import BELOW_PANEL_FIELDS, STRUCTURED_COLUMNS
from .slide_verified import ALL_SLIDE_FIELDS, VERIFIED_SLIDE_BY_ORG

PILOT_BY_ORG = VERIFIED_SLIDE_BY_ORG

# Spreadsheet header may differ from slide field names
COLUMN_ALIASES: dict[str, str] = {
    "Service lines": "Platforms",
}

SECTION_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FILL = PatternFill(start_color="0D1526", end_color="0D1526", fill_type="solid")
HEADER_FONT = Font(bold=True, color="94A3B8", size=10)
LINK_FONT = Font(color="0563C1", underline="single")


def _header_map(ws) -> dict[str, int]:
    m: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(2, col).value
        if v:
            key = str(v).strip()
            m[key] = col
            # SheetJS / dashboard use raw row-2 labels (e.g. leading space on G/Y/R)
            if key != str(v):
                m[str(v)] = col
    return m


def _col_for_field(col_index: dict[str, int], field: str) -> int | None:
    if field in col_index:
        return col_index[field]
    stripped = field.strip()
    if stripped in col_index:
        return col_index[stripped]
    return None


def _ensure_columns(ws, col_index: dict[str, int]) -> dict[str, int]:
    """Add any missing structured columns (incl. Salesforce URL)."""
    start = max(col_index.values(), default=2) + 1
    section_for: dict[str, str] = {
        "Salesforce URL": "Subscription Review",
        "AAR $": "Subscription Review",
        "Collab AE/SE": "Subscription Review",
        "Service lines": "Subscription Review",
        "Platforms": "Subscription Review",
        "Sub term (months)": "Subscription Review",
        "Subscription dates": "Subscription Review",
        "Trend active users 90d": "Subscription Review",
        "Trend call volume 90d": "Subscription Review",
        "Lic Professional (used/entitled)": "Provisioning & Usage",
        "Lic Standard (used/entitled)": "Provisioning & Usage",
        "Lic Workspace (used/entitled)": "Provisioning & Usage",
        "Active % of provisioned": "Provisioning & Usage",
        "External calls": "Provisioning & Usage",
        "Meetings usage": "Provisioning & Usage",
        "Messaging usage": "Provisioning & Usage",
        "Auto Attendant count": "Feature Use & Add-ons",
        "Hunt Groups count": "Feature Use & Add-ons",
        "Call Queues count": "Feature Use & Add-ons",
        "Connected-UC (Y/N)": "Feature Use & Add-ons",
        "Virtual Lines count": "Feature Use & Add-ons",
        "Data gathered by": "Business Insight",
        "Data gathered date": "Business Insight",
    }
    seen_sections: set[str] = set()
    for name in STRUCTURED_COLUMNS:
        if name in col_index:
            continue
        section = section_for.get(name, "Business Insight")
        ws.cell(1, start, value=section if section not in seen_sections else "")
        seen_sections.add(section)
        ws.cell(2, start, value=name)
        ws.cell(1, start).fill = SECTION_FILL
        ws.cell(2, start).fill = HEADER_FILL
        ws.cell(2, start).font = HEADER_FONT
        ws.cell(2, start).alignment = Alignment(wrap_text=True, vertical="top")
        col_index[name] = start
        start += 1
    return col_index


def clear_below_panel_columns(ws, col_index: dict[str, int]) -> int:
    """Empty KPI strip / bottom-note columns (not on the three slide panels)."""
    cleared = 0
    for field in BELOW_PANEL_FIELDS:
        col = _col_for_field(col_index, field)
        if not col:
            continue
        for row in range(3, ws.max_row + 1):
            if ws.cell(row, col).value not in (None, ""):
                ws.cell(row, col).value = None
                cleared += 1
    return cleared


def _sync_salesforce_urls(ws, col_index: dict[str, int]) -> int:
    link_col = col_index.get("Opportunity (linked)")
    url_col = col_index.get("Salesforce URL")
    if not link_col or not url_col:
        return 0
    n = 0
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row, link_col)
        url = None
        if cell.hyperlink and cell.hyperlink.target:
            url = cell.hyperlink.target
        elif cell.value and str(cell.value).startswith("http"):
            url = str(cell.value)
        if url:
            ws.cell(row, url_col, value=url)
            cell.font = LINK_FONT
            if not cell.hyperlink:
                cell.hyperlink = url
                cell.value = cell.value if cell.value and not str(cell.value).startswith("http") else "Salesforce Link"
            n += 1
    return n


def _norm_money(val: Any) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("$", "").strip().upper()
    mult = 1.0
    if s.endswith("K"):
        mult = 1_000.0
        s = s[:-1]
    elif s.endswith("M"):
        mult = 1_000_000.0
        s = s[:-1]
    try:
        return float(s) * mult
    except ValueError:
        return None


def _values_match(field: str, expected: Any, actual: Any) -> bool:
    if expected in (None, "") and actual in (None, ""):
        return True
    if field in ("TCV $", "AAR $"):
        e, a = _norm_money(expected), _norm_money(actual)
        if e is None or a is None:
            return str(expected).strip() == str(actual).strip()
        return abs(e - a) < max(1.0, abs(e) * 0.001)
    return str(expected).strip() == str(actual).strip()


def verify_slide_sync(path: Path) -> list[dict[str, Any]]:
    """Return list of field mismatches for orgs with verified slides."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    col_index = _header_map(ws)
    mismatches: list[dict[str, Any]] = []

    for row in range(3, ws.max_row + 1):
        row_data = {h: ws.cell(row, col_index[h]).value for h in col_index}
        org = str(row_data.get("Customer org id") or "").strip().lower()
        slide = VERIFIED_SLIDE_BY_ORG.get(org)
        if not slide:
            continue
        name = row_data.get("Opportunity Name") or org
        for field, expected in slide.items():
            target = COLUMN_ALIASES.get(field, field)
            col = _col_for_field(col_index, target)
            if col is None:
                continue  # column not on workbook — user may have removed it
            actual = ws.cell(row, col).value
            if not _values_match(field, expected, actual):
                mismatches.append(
                    {
                        "row": row,
                        "customer": str(name)[:60],
                        "org_id": org,
                        "field": field,
                        "slide": expected,
                        "spreadsheet": actual,
                    }
                )
    wb.close()
    return mismatches


def rebuild_powerpoint_columns(
    path: Path, backup: bool = True, *, ensure_columns: bool = False
) -> dict[str, Any]:
    path = Path(path)
    if backup:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(path, path.with_name(path.stem + f".backup_{ts}" + path.suffix))

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_index = _header_map(ws)
    if ensure_columns:
        col_index = _ensure_columns(ws, col_index)

    struct_cols = [c for c in STRUCTURED_COLUMNS if c in col_index]
    slide_fields = [f for f in ALL_SLIDE_FIELDS if _col_for_field(col_index, f)]

    filled = 0
    slide_rows = 0
    for row in range(3, ws.max_row + 1):
        row_data = {h: ws.cell(row, col_index[h]).value for h in col_index}
        org = str(row_data.get("Customer org id") or "").strip().lower()
        slide = VERIFIED_SLIDE_BY_ORG.get(org)

        if slide:
            slide_rows += 1
            merged = dict(slide)
        else:
            # Non-slide rows: conservative parse from notes only
            analytics = str(row_data.get("Notes from Calling Analytics") or "")
            features = str(row_data.get("Notes from provisioned features") or "")
            addons = str(row_data.get("Add-ons included or not") or "")
            merged = parse_notes_to_structured(analytics, features, addons)
            pct = compute_active_pct(
                row_data.get("Providioned Lic Calling"),
                row_data.get("Active Lic Calling"),
            )
            if pct:
                merged["Active % of provisioned"] = pct

        for field, val in merged.items():
            target = COLUMN_ALIASES.get(field, field)
            col = _col_for_field(col_index, target)
            if col is None or val in (None, ""):
                continue
            ws.cell(row, col, value=val)
            filled += 1

    urls = _sync_salesforce_urls(ws, col_index)
    below_cleared = clear_below_panel_columns(ws, col_index)
    wb.save(path)
    mismatches = verify_slide_sync(path)
    return {
        "path": str(path),
        "structured_columns": struct_cols,
        "cells_filled": filled,
        "salesforce_urls_synced": urls,
        "below_panel_cells_cleared": below_cleared,
        "slide_verified_rows": slide_rows,
        "slide_mismatches_after_sync": mismatches,
    }


if __name__ == "__main__":
    import sys

    target = (
        Path(sys.argv[1])
        if len(sys.argv) > 1
        else Path.home() / "Downloads/Check Back Initiative Customer Analysis.xlsx"
    )
    print(rebuild_powerpoint_columns(target))

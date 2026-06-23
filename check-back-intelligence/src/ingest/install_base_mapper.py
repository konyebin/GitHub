"""Map Install Base report rows to Check Back Initiative format."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import openpyxl
import yaml

CONFIG_PATH = Path(__file__).resolve().parents[2] / "config" / "field_mapping.yaml"


def load_config() -> dict[str, Any]:
    with CONFIG_PATH.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def _first_number(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    s = str(val).replace(",", "").replace("$", "")
    parts = s.split()
    for p in parts:
        try:
            return float(p)
        except ValueError:
            continue
    try:
        return float(s)
    except ValueError:
        return 0.0


def _format_license(mt: Any, di: Any) -> str:
    parts = []
    if mt not in (None, ""):
        parts.append(f"MT: {int(_first_number(mt))}")
    if di not in (None, ""):
        parts.append(f"DI: {int(_first_number(di))}")
    return " ".join(parts) if parts else ""


def map_row(ib_row: dict[str, Any], config: dict[str, Any] | None = None) -> tuple[dict[str, str], list[dict[str, str]]]:
    """Return Check Back row dict and gap report entries."""
    config = config or load_config()
    mapping = config.get("install_base_to_checkback", {})
    risk_map = config.get("risk_to_gyr", {})
    review = config.get("review_flags", {})
    gaps: list[dict[str, str]] = []
    out: dict[str, str] = {}

    for ib_col, cb_col in mapping.items():
        val = ib_row.get(ib_col, "")
        if ib_col == "Risk2_0_current":
            val = risk_map.get(str(val), str(val)[:1].upper() if val else "")
        elif ib_col in ("Webex Calling MT Provisioned Seats", "Webex Calling DI Provisioned Seats"):
            if cb_col == "Providioned Lic Calling" and "Providioned Lic Calling" not in out:
                val = _format_license(
                    ib_row.get("Webex Calling MT Provisioned Seats"),
                    ib_row.get("Webex Calling DI Provisioned Seats"),
                )
            else:
                continue
        elif ib_col == "Cloud Calling Billed Seats" and cb_col == "Entitled Lic Calling":
            seats = _first_number(val) or _first_number(ib_row.get("Total Billed Seats"))
            val = f"{int(seats)} workspace" if seats else ""
        if val not in (None, ""):
            out[cb_col] = str(val) if not isinstance(val, (int, float)) else val
        if ib_col in review and ib_col in ib_row and ib_row[ib_col]:
            gaps.append(
                {
                    "field": cb_col,
                    "source": ib_col,
                    "confidence": "medium",
                    "suggested_action": f"Verify mapping: {ib_col} → {review[ib_col]}",
                }
            )

    for ib_col in config.get("check_back_columns", []):
        if ib_col not in out.values() and ib_col not in out:
            src_used = [k for k, v in mapping.items() if v == ib_col]
            if not src_used:
                gaps.append(
                    {
                        "field": ib_col,
                        "source": "install_base",
                        "confidence": "low",
                        "suggested_action": "No install base source; use PDF or manual entry",
                    }
                )

    return out, gaps


def map_workbook(
    input_path: str | Path,
    output_json: str | Path | None = None,
    gap_csv: str | Path | None = None,
    limit: int | None = None,
) -> list[dict[str, str]]:
    """Map all rows from install base xlsx; optionally write JSON + gap CSV."""
    config = load_config()
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, str]] = []
    all_gaps: list[dict[str, str]] = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if limit and i > limit + 1:
            break
        ib_row = dict(zip(headers, row))
        if not any(v not in (None, "") for v in ib_row.values()):
            continue
        cb_row, gaps = map_row(ib_row, config)
        if cb_row.get("Opportunity Name") or cb_row.get("Partner"):
            rows.append(cb_row)
            for g in gaps:
                g["row"] = str(cb_row.get("Opportunity Name", i))
                all_gaps.append(g)

    wb.close()

    if output_json:
        Path(output_json).parent.mkdir(parents=True, exist_ok=True)
        Path(output_json).write_text(json.dumps(rows, indent=2, default=str), encoding="utf-8")

    if gap_csv:
        Path(gap_csv).parent.mkdir(parents=True, exist_ok=True)
        with Path(gap_csv).open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["row", "field", "source", "confidence", "suggested_action"])
            w.writeheader()
            w.writerows(all_gaps)

    return rows

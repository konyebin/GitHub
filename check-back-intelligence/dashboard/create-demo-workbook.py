#!/usr/bin/env python3
"""Create a sanitized demo Check Back workbook for public GitHub Pages hosting."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = [
    "Opportunity Name",
    "Opportunity (linked)",
    "SL2",
    "TCV $",
    "Closed on (MM/YY)",
    "Competitor",
    "Migrating from",
    "Migrating to",
    "Partner",
    "CSM Engagement Model (linked)",
    "CSM name",
    "Sub #",
    "Sub Term",
    "Add-ons included or not",
    "Calling Setup Assist included (Y/N)",
    "Entitled Lic Calling",
    "Providioned Lic Calling",
    "Active Lic Calling",
    "Customer org id",
    "Notes from Calling Analytics",
    "Notes from provisioned features",
    "CCEP trial (Y/N)",
    " (G/Y/R)",
    "TAC/BEMS",
]

SECTION = "Business Insight"
FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HDR_FILL = PatternFill(start_color="0D1526", end_color="0D1526", fill_type="solid")
HDR_FONT = Font(bold=True, color="94A3B8", size=10)

DEMO_ROWS = [
    {
        "Opportunity Name": "Demo School District (sample)",
        "SL2": "US PS Market",
        "TCV $": 850000,
        "Partner": "Demo Partner LLC",
        "Sub #": "Sub900001",
        "Entitled Lic Calling": "PL: 4200, WS: 1750",
        "Providioned Lic Calling": "PL: 283, WS: 1200",
        "Active Lic Calling": "1180 active",
        " (G/Y/R)": "G",
        "Notes from Calling Analytics": "Sample row — upload your own Check Back workbook to analyze real accounts.",
    },
    {
        "Opportunity Name": "Demo Hospital System (sample)",
        "SL2": "USPS",
        "TCV $": 1200000,
        "Partner": "Example Integrator",
        "Sub #": "Sub900002",
        "Entitled Lic Calling": "PL: 1860, WS: 775",
        "Providioned Lic Calling": "PL: 1326, WS: 374",
        "Active Lic Calling": "1300",
        " (G/Y/R)": "Y",
        "Notes from Calling Analytics": "Workspace under-utilized; Professional near target.",
    },
    {
        "Opportunity Name": "Demo County Government (sample)",
        "SL2": "US SLED",
        "TCV $": 640000,
        "Partner": "Public Sector Partner",
        "Sub #": "Sub900003",
        "Entitled Lic Calling": "PL: 1800, WS: 750",
        "Providioned Lic Calling": "PL: 129, WS: 57",
        "Active Lic Calling": "158",
        " (G/Y/R)": "R",
        "Notes from Calling Analytics": "Low adoption vs entitled licenses.",
    },
]


def build_demo_workbook(dest: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col, name in enumerate(HEADERS, start=1):
        ws.cell(1, col, value=SECTION if col == 1 else "")
        ws.cell(1, col).fill = FILL
        ws.cell(2, col, value=name)
        ws.cell(2, col).fill = HDR_FILL
        ws.cell(2, col).font = HDR_FONT
        ws.cell(2, col).alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, row_data in enumerate(DEMO_ROWS, start=3):
        for col, name in enumerate(HEADERS, start=1):
            val = row_data.get(name)
            if val is not None:
                ws.cell(row_idx, col, value=val)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("check_back_default.xlsx")
    build_demo_workbook(out)
    print(f"Wrote demo workbook → {out}")

"""Check Back dashboard orchestration for the Webex bot."""

from __future__ import annotations

import re
import subprocess
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from .bot_tunnel import build_dashboard_link, dashboard_base_url

ROOT = Path(__file__).resolve().parent.parent
BASELINE_XLSX = ROOT / "output" / "Check_Back_standardized.xlsx"
SYNC_SCRIPT = ROOT / "dashboard" / "sync-default-workbook.sh"
GYR_COL = " (G/Y/R)"
LICENSE_COL = "Provisioned/Entitled Lic Calling"
PAIR_RE = re.compile(
    r"\b(?:PL|WS|Professional|Workspace)\s*:?\s*(\d[\d,]*)\s*/\s*(\d[\d,]*)",
    re.I,
)


@dataclass
class BotReply:
    markdown: str
    dashboard_url: str | None = None


def _cell_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def sync_baseline() -> Path:
    """Copy output/Check_Back_standardized.xlsx → dashboard/check_back_default.xlsx."""
    if not SYNC_SCRIPT.is_file():
        raise FileNotFoundError(f"Missing sync script: {SYNC_SCRIPT}")
    if not BASELINE_XLSX.is_file():
        raise FileNotFoundError(
            f"Baseline workbook not found: {BASELINE_XLSX}\n"
            "Edit that file or set CHECK_BACK_XLSX before running the bot."
        )
    subprocess.run(["/bin/bash", str(SYNC_SCRIPT)], check=True, cwd=str(SYNC_SCRIPT.parent))
    dest = ROOT / "dashboard" / "check_back_default.xlsx"
    if not dest.is_file():
        raise FileNotFoundError(f"Sync did not produce {dest}")
    return dest


def _detect_header_row(ws) -> int:
    for r in range(1, min(6, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            if _cell_str(ws.cell(r, c).value) == "Opportunity Name":
                return r
    return 2


def _headers(ws, header_row: int) -> dict[str, int]:
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = _cell_str(ws.cell(header_row, c).value)
        if h:
            out[h] = c
    return out


def _license_totals(cell: str) -> tuple[float, float]:
    prov = ent = 0.0
    for m in PAIR_RE.finditer(cell or ""):
        prov += float(m.group(1).replace(",", ""))
        ent += float(m.group(2).replace(",", ""))
    if prov or ent:
        return ent, prov
    nums = re.findall(r"[\d,]+", cell or "")
    if len(nums) == 1:
        n = float(nums[0].replace(",", ""))
        return n, n
    return 0.0, 0.0


def load_portfolio_stats(path: Path | None = None) -> dict:
    wb_path = path or BASELINE_XLSX
    if not wb_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {wb_path}")

    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb.active
    header_row = _detect_header_row(ws)
    cols = _headers(ws, header_row)

    name_col = cols.get("Opportunity Name", 1)
    gyr_col = cols.get(GYR_COL)
    lic_col = cols.get(LICENSE_COL)
    act_col = cols.get("Active Lic Calling")
    tcv_col = cols.get("TCV $")

    rows: list[dict] = []
    gyr_counts: dict[str, int] = {"G": 0, "Y": 0, "R": 0}
    total_ent = total_prov = total_act = 0.0

    for r in range(header_row + 1, ws.max_row + 1):
        name = _cell_str(ws.cell(r, name_col).value)
        if not name:
            continue
        lic_raw = _cell_str(ws.cell(r, lic_col).value) if lic_col else ""
        act_raw = ws.cell(r, act_col).value if act_col else ""
        ent, prov = _license_totals(lic_raw)
        act = float(_cell_str(act_raw).replace(",", "") or 0) if act_raw not in (None, "") else 0.0
        gyr = _cell_str(ws.cell(r, gyr_col).value).upper()[:1] if gyr_col else ""
        if gyr in gyr_counts:
            gyr_counts[gyr] += 1
        total_ent += ent
        total_prov += prov
        total_act += act
        rows.append(
            {
                "name": name,
                "gyr": gyr or "—",
                "entitled": ent,
                "provisioned": prov,
                "active": act,
                "tcv": _cell_str(ws.cell(r, tcv_col).value) if tcv_col else "",
            }
        )

    wb.close()
    prov_ent_pct = round(100 * total_prov / total_ent) if total_ent else 0
    act_prov_pct = round(100 * total_act / total_prov) if total_prov else 0
    return {
        "path": str(wb_path),
        "updated": datetime.fromtimestamp(wb_path.stat().st_mtime, tz=timezone.utc).strftime(
            "%Y-%m-%d %H:%M UTC"
        ),
        "count": len(rows),
        "gyr": gyr_counts,
        "total_entitled": int(total_ent),
        "total_provisioned": int(total_prov),
        "total_active": int(total_act),
        "prov_ent_pct": prov_ent_pct,
        "act_prov_pct": act_prov_pct,
        "rows": rows,
    }


def format_summary_markdown(stats: dict, *, dashboard_url: str) -> str:
    g = stats["gyr"]
    lines = [
        "## Check Back Portfolio Summary",
        f"**Baseline:** `{Path(stats['path']).name}`",
        f"**Updated:** {stats['updated']}",
        f"**Opportunities:** {stats['count']}",
        "",
        "| Metric | Value |",
        "|--------|-------|",
        f"| Provisioned / Entitled | {stats['prov_ent_pct']}% ({stats['total_provisioned']:,} / {stats['total_entitled']:,}) |",
        f"| Active / Provisioned | {stats['act_prov_pct']}% ({stats['total_active']:,} / {stats['total_provisioned']:,}) |",
        f"| G / Y / R | {g.get('G', 0)} / {g.get('Y', 0)} / {g.get('R', 0)} |",
        "",
        "**Top accounts (by entitled licenses)**",
    ]
    top = sorted(stats["rows"], key=lambda x: x["entitled"], reverse=True)[:8]
    for row in top:
        short = row["name"][:52] + ("…" if len(row["name"]) > 52 else "")
        lines.append(
            f"- **{short}** — {row['gyr']} · "
            f"{int(row['provisioned']):,}/{int(row['entitled']):,} prov/ent · "
            f"{int(row['active']):,} active"
        )
    lines.extend(
        [
            "",
            f"**[Open Check Back dashboard]({dashboard_url})**",
            "",
            "_Upload or sync the baseline xlsx in the dashboard to refresh charts._",
        ]
    )
    return "\n".join(lines)


def help_markdown() -> str:
    base = dashboard_base_url()
    link = f"\nDashboard tunnel: `{base}`" if base else "\nStart with `bash webex_bot/run-check-back-bot.sh`."
    return (
        "## Check Back Bot commands\n"
        "- `dashboard` — link to the Check Back dashboard\n"
        "- `summary` / `portfolio` — KPI snapshot from `output/Check_Back_standardized.xlsx`\n"
        "- `refresh` / `sync baseline` — re-sync baseline into the dashboard, then post link + summary\n"
        "- `help` — this message\n"
        + link
    )


def dashboard_link_reply() -> BotReply:
    url = build_dashboard_link()
    md = (
        "## Check Back Intelligence Dashboard\n\n"
        f"**[Open dashboard]({url})**\n\n"
        f"Baseline workbook: `output/Check_Back_standardized.xlsx`"
    )
    return BotReply(markdown=md, dashboard_url=url)


def summary_reply() -> BotReply:
    stats = load_portfolio_stats()
    url = build_dashboard_link()
    return BotReply(markdown=format_summary_markdown(stats, dashboard_url=url), dashboard_url=url)


def refresh_reply() -> BotReply:
    sync_baseline()
    return summary_reply()
